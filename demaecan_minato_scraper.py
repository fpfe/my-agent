#!/usr/bin/env python3
"""
出前館 — 港区 カテゴリ別レストラン調査
Playwright で白金1丁目(港区中心)の配達エリアを巡回し、
ジャンル別の店舗数・店名一覧を取得する。

Output: demaecan_minato_categories.xlsx
"""
from __future__ import annotations
import re, time, sys
from collections import defaultdict
from datetime import datetime
from typing import Optional

MISSING = []
try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
except ImportError:
    MISSING.append("playwright")
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    MISSING.append("openpyxl")
try:
    from bs4 import BeautifulSoup
except ImportError:
    MISSING.append("beautifulsoup4")

if MISSING:
    print(f"[ERROR] Missing: {', '.join(MISSING)} — pip install {' '.join(MISSING)}")
    sys.exit(1)

# ── config ────────────────────────────────────────────────────────────────────
# 港区白金1丁目 (代表住所コード)
AREA_CODE = "13103014001"
BASE_URL   = f"https://demae-can.com/search/delivery/{AREA_CODE}"
OUTPUT_FILE = "demaecan_minato_categories.xlsx"

# 出前館 ジャンルコード一覧
GENRES = [
    ("弁当",               "01"),
    ("ピザ",               "02"),
    ("寿司・海鮮",          "03"),
    ("中華",               "04"),
    ("そば・うどん",        "06"),
    ("和食",               "07"),
    ("とんかつ・天ぷら",     "08"),
    ("たこ焼き・お好み焼き", "09"),
    ("焼肉・しゃぶしゃぶ",   "10"),
    ("カレーライス",        "11"),
    ("スイーツ",            "12"),
    ("お酒",               "15"),
    ("バーガー・サンドイッチ","18"),
    ("コンビニ",           "19"),
    ("洋食",               "50"),
    ("パーティー・ケータリング","55"),
    ("エスニック",          "60"),
    ("ラーメン",            "61"),
    ("丼",                 "62"),
    ("パスタ・イタリアン",   "67"),
    ("インドカレー",        "112"),
    ("ファミレス",          "96"),
    ("韓国料理",            "97"),
    ("ファストフード",       "98"),
    ("フライドチキン・唐揚げ","45"),
    ("ハンバーグ・ステーキ", "051"),
    ("ヘルシー料理",        "52"),
    ("カフェ・ドリンク",     "59"),
    ("お店価格",            "100"),
]


def log(msg: str):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")


def extract_shops(html: str) -> list[dict]:
    """
    Parse shop cards from page HTML.
    Returns list of dicts: {shop_id, name, rating, wait_min, delivery_fee}
    """
    soup = BeautifulSoup(html, "lxml")
    shops = []
    seen_ids: set[str] = set()

    for a in soup.find_all("a", href=True):
        href = a["href"]
        m = re.match(r"/shop/menu/(\d+)", href)
        if not m:
            continue
        shop_id = m.group(1)
        if shop_id in seen_ids:
            continue
        seen_ids.add(shop_id)

        # Get parent article for full info
        article = a.find_parent("article")
        full_text = article.get_text(" ", strip=True) if article else a.get_text(strip=True)

        # Extract shop name: first <p> or <h2>/<h3> in article
        name = ""
        if article:
            for tag in article.find_all(["h2", "h3", "p", "span"]):
                t = tag.get_text(strip=True)
                # Skip badges / labels
                if t and len(t) > 3 and not re.match(r'^[\d.]+$', t) and "送料" not in t and "クーポン" not in t and "分" not in t and "対象" not in t:
                    name = t
                    break
        if not name:
            # fallback: strip digits/badges from full text
            name = re.sub(r'(お店価格|送料無料|クーポン|対象|[\d.]+分|送料\d+円|標準送料\d+円)', '', full_text).strip()[:60]

        # Rating
        rating_m = re.search(r'(\d\.\d)', full_text)
        rating = float(rating_m.group(1)) if rating_m else None

        # Wait time
        wait_m = re.search(r'(\d+)分', full_text)
        wait_min = int(wait_m.group(1)) if wait_m else None

        # Delivery fee
        fee_m = re.search(r'送料(\d+)円', full_text)
        delivery_fee = int(fee_m.group(1)) if fee_m else (0 if "送料無料" in full_text else None)

        shops.append({
            "shop_id": shop_id,
            "name": name,
            "rating": rating,
            "wait_min": wait_min,
            "delivery_fee": delivery_fee,
        })

    return shops


def scroll_and_collect(page, max_scrolls: int = 30) -> list[dict]:
    """Scroll to load lazy-loaded shops, return deduplicated list."""
    seen_ids: set[str] = set()
    all_shops: list[dict] = []
    stale = 0
    prev_count = 0

    for _ in range(max_scrolls):
        shops = extract_shops(page.content())
        for s in shops:
            if s["shop_id"] not in seen_ids:
                seen_ids.add(s["shop_id"])
                all_shops.append(s)

        page.evaluate("window.scrollBy(0, 1200)")
        time.sleep(1.5)

        if len(all_shops) == prev_count:
            stale += 1
            if stale >= 4:
                break
        else:
            stale = 0
        prev_count = len(all_shops)

    return all_shops


def scrape_genre(page, genre_label: str, genre_code: str) -> list[dict]:
    url = f"{BASE_URL}/{genre_code}"
    log(f"  [{genre_label}] {url}")
    try:
        page.goto(url, timeout=25000, wait_until="domcontentloaded")
    except PWTimeout:
        log(f"    [!] Timeout navigating to {genre_label}")

    try:
        page.wait_for_selector("article", timeout=12000)
    except PWTimeout:
        log(f"    [!] No shop cards for {genre_label}")
        return []

    time.sleep(2)
    shops = scroll_and_collect(page, max_scrolls=40)
    log(f"    → {len(shops)} 件")
    return shops


def run_all() -> tuple[dict[str, list[dict]], list[dict]]:
    genre_results: dict[str, list[dict]] = {}
    all_shops: list[dict] = []

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=False,
            args=["--disable-http2"],
        )
        ctx = browser.new_context(
            locale="ja-JP",
            timezone_id="Asia/Tokyo",
            viewport={"width": 1280, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
        )
        page = ctx.new_page()

        # ── Step 1: Total (all genres) ─────────────────────────────────────
        log("全ジャンル(合計)取得中...")
        try:
            page.goto(BASE_URL, timeout=25000, wait_until="domcontentloaded")
            page.wait_for_selector("article", timeout=15000)
        except PWTimeout:
            log("  [!] 全ジャンルページ タイムアウト")

        time.sleep(3)
        all_shops = scroll_and_collect(page, max_scrolls=60)
        log(f"  → 全店舗合計: {len(all_shops)} 件")

        # ── Step 2: Per genre ──────────────────────────────────────────────
        log(f"\n{len(GENRES)} ジャンルを順次取得...")
        for genre_label, genre_code in GENRES:
            shops = scrape_genre(page, genre_label, genre_code)
            genre_results[genre_label] = shops
            time.sleep(2)

        browser.close()

    return genre_results, all_shops


# ── Excel export ──────────────────────────────────────────────────────────────

def export_excel(genre_results: dict[str, list[dict]], all_shops: list[dict]):
    wb = openpyxl.Workbook()
    RED    = "DA3734"   # 出前館 brand red
    HDR_FILL  = PatternFill("solid", fgColor=RED)
    HDR_FONT  = Font(bold=True, color="FFFFFF")
    HDR_ALIGN = Alignment(horizontal="center")
    ZEBRA     = PatternFill("solid", fgColor="FFF0F0")

    def style_header(ws, row_num: int):
        for cell in ws[row_num]:
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = HDR_ALIGN

    # ── Sheet 1: Genre summary ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "ジャンル別件数"
    ws1["A1"] = "出前館 港区 — ジャンル別レストラン数"
    ws1["A1"].font = Font(bold=True, size=14, color=RED)
    ws1["A2"] = f"取得日時: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws1["A3"] = f"配達エリア: 東京都港区白金1丁目 (代表コード: {AREA_CODE})"
    ws1["A4"] = f"全店舗合計: {len(all_shops)} 件"
    ws1.append([])
    ws1.append(["ジャンル", "件数", "ジャンルコード"])
    style_header(ws1, 6)

    sorted_genres = sorted(genre_results.items(), key=lambda x: -len(x[1]))
    for i, (label, shops) in enumerate(sorted_genres, 7):
        # find code
        code = next((c for l, c in GENRES if l == label), "")
        ws1.append([label, len(shops), code])
        if i % 2 == 0:
            for cell in ws1[i]:
                cell.fill = ZEBRA

    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 8
    ws1.column_dimensions["C"].width = 14
    ws1.freeze_panes = "A7"

    # ── Sheet 2: All shops ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("全店舗リスト")
    ws2.append(["#", "店舗名", "評価", "到着時間(分)", "配送料(円)", "店舗ID"])
    style_header(ws2, 1)
    for i, s in enumerate(all_shops, 1):
        ws2.append([
            i, s["name"], s["rating"],
            s["wait_min"], s["delivery_fee"], s["shop_id"]
        ])
        if i % 2 == 0:
            for col in range(1, 7):
                ws2.cell(i + 1, col).fill = ZEBRA
    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 50
    ws2.column_dimensions["C"].width = 8
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 14
    ws2.column_dimensions["F"].width = 12
    ws2.freeze_panes = "A2"

    # ── Per-genre sheets ───────────────────────────────────────────────────
    for label, shops in sorted_genres:
        safe_title = re.sub(r'[\\/*?:\[\]]', '', label)[:28]
        ws = wb.create_sheet(safe_title)
        ws.append(["#", f"{label} — {len(shops)} 件", "評価", "到着(分)", "配送料(円)"])
        style_header(ws, 1)
        for i, s in enumerate(shops, 1):
            ws.append([i, s["name"], s["rating"], s["wait_min"], s["delivery_fee"]])
            if i % 2 == 0:
                for col in range(1, 6):
                    ws.cell(i + 1, col).fill = ZEBRA
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 12
        ws.freeze_panes = "A2"

    wb.save(OUTPUT_FILE)
    log(f"Saved → {OUTPUT_FILE}")


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    print("\n" + "█" * 60)
    print("  出前館 — 港区 ジャンル別レストラン調査")
    print("█" * 60)
    print(f"  開始: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

    genre_results, all_shops = run_all()

    print(f"\n{'='*60}")
    print("  結果サマリー")
    print(f"{'='*60}")
    print(f"  全店舗合計: {len(all_shops)} 件\n")

    sorted_genres = sorted(genre_results.items(), key=lambda x: -len(x[1]))
    for label, shops in sorted_genres:
        bar = "█" * min(len(shops) // 2, 40)
        print(f"  {label:<25} {len(shops):>4} 件  {bar}")

    export_excel(genre_results, all_shops)

    print(f"\n  完了: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  出力: {OUTPUT_FILE}\n")


if __name__ == "__main__":
    main()
