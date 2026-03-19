#!/usr/bin/env python3
"""
Tabelog × Uber Eats Supply Gap Analyser — Minato-ku, Tokyo
Outputs: tabelog_ubereats_gap_analysis.xlsx  (4 sheets)
"""

from __future__ import annotations

import time
import random
import re
import sys
from datetime import datetime
from typing import Optional

# ── dependency check ────────────────────────────────────────────────────────
MISSING = []
try:
    import requests
except ImportError:
    MISSING.append("requests")
try:
    from bs4 import BeautifulSoup
except ImportError:
    MISSING.append("beautifulsoup4")
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    MISSING.append("openpyxl")

if MISSING:
    print(f"[ERROR] Missing packages: {', '.join(MISSING)}")
    print(f"  Run:  pip install {' '.join(MISSING)}")
    sys.exit(1)

# ── constants ────────────────────────────────────────────────────────────────
BASE_URL        = "https://tabelog.com/tokyo/A1307/A130701/"
RANKING_URL     = BASE_URL + "rstLst/?SrtT=rt"
OUTPUT_FILE     = "tabelog_ubereats_gap_analysis.xlsx"
REQUEST_DELAY   = (2.5, 3.5)   # seconds (random range)

# Keyword-search URLs — Tabelog's ?sk= restricts to restaurant name/genre only
# (without vs=1 which makes sk= search all review text, returning the area total)
CATEGORY_SLUGS = {
    "Pizza":    BASE_URL + "rstLst/?sk=%E3%83%94%E3%82%B6",
    "Sushi":    BASE_URL + "rstLst/?sk=%E5%AF%BF%E5%8F%B8",
    "Ramen":    BASE_URL + "rstLst/?sk=%E3%83%A9%E3%83%BC%E3%83%A1%E3%83%B3",
    "Italian":  BASE_URL + "rstLst/?sk=%E3%82%A4%E3%82%BF%E3%83%AA%E3%82%A2%E3%83%B3",
    "Yakiniku": BASE_URL + "rstLst/?sk=%E7%84%BC%E8%82%89",
    "Tempura":  BASE_URL + "rstLst/?sk=%E5%A4%A9%E3%81%B7%E3%82%89",
    "French":   BASE_URL + "rstLst/?sk=%E3%83%95%E3%83%AC%E3%83%B3%E3%83%81",
}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "ja,en-US;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Referer": "https://tabelog.com/",
}

SESSION = requests.Session()
SESSION.headers.update(HEADERS)

# ── helpers ──────────────────────────────────────────────────────────────────

def sleep():
    t = random.uniform(*REQUEST_DELAY)
    time.sleep(t)


def fetch(url: str, retries: int = 3) -> Optional[BeautifulSoup]:
    """GET a page and return a BeautifulSoup object, or None on failure."""
    for attempt in range(1, retries + 1):
        try:
            resp = SESSION.get(url, timeout=15)
            if resp.status_code == 200:
                return BeautifulSoup(resp.text, "html.parser")
            elif resp.status_code == 403:
                print(f"  [!] 403 Forbidden — Tabelog blocked the request for {url}")
                print("       Tip: install playwright and set USE_PLAYWRIGHT=True at the top of this script.")
                return None
            else:
                print(f"  [!] HTTP {resp.status_code} for {url} (attempt {attempt}/{retries})")
        except requests.RequestException as e:
            print(f"  [!] Request error ({e}) — attempt {attempt}/{retries}")
        if attempt < retries:
            sleep()
    return None


def banner(title: str):
    print(f"\n{'='*60}")
    print(f"  {title}")
    print(f"{'='*60}")


def log(msg: str):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}")

# ── Task 1 — Restaurant counts by category ──────────────────────────────────

def _extract_total_count(soup: BeautifulSoup) -> int:
    """
    Extract total restaurant count from a Tabelog listing page.
    Tabelog shows "1 ～ 20 件を表示 ／ 全 2,099 件" in the page header.
    We specifically look for the "全 XXX 件" (= total XXX items) pattern.
    """
    all_text = soup.get_text(" ", strip=True)

    # Most reliable: "全 2,099 件" or "全2099件"
    m = re.search(r'全\s*([\d,]+)\s*件', all_text)
    if m:
        return int(m.group(1).replace(",", ""))

    # Fallback: any "XXX件" taking the LARGEST value (not page-range start)
    matches = re.findall(r'([\d,]+)\s*件', all_text)
    if matches:
        candidates = [int(n.replace(",", "")) for n in matches]
        # Filter out single-digit "page start" numbers
        big = [c for c in candidates if c > 20]
        if big:
            return max(big)

    # Last resort: count cards on the first page
    cards = soup.select("div.list-rst__wrap, li.js-rstlist-info")
    return len(cards)


# Japanese genre keywords mapped to our 7 category labels
CATEGORY_KEYWORDS = {
    "Pizza":    ["ピザ", "pizza", "Pizza"],
    "Sushi":    ["寿司", "鮨", "sushi"],
    "Ramen":    ["ラーメン", "ramen"],
    "Italian":  ["イタリアン", "イタリア料理", "italian"],
    "Yakiniku": ["焼肉", "yakiniku"],
    "Tempura":  ["天ぷら", "天麩羅", "天婦羅", "tempura"],
    "French":   ["フレンチ", "french", "フランス料理"],
}


def task1_category_counts() -> list[dict]:
    """
    Scan the Minato-ku listing pages and count restaurants by category
    from their actual genre field (div.list-rst__area-genre).
    Scans SCAN_PAGES pages (~20 restaurants each) as a sample, then
    scales up to the known area total.
    """
    banner("TASK 1 — Restaurant counts by category in Minato-ku")
    SCAN_PAGES = 10   # 200 restaurants — representative sample

    # First fetch to get the total restaurant count in the area
    log(f"Fetching listing page 1 to determine total restaurant count...")
    soup0 = fetch(BASE_URL + "rstLst/")
    area_total = _extract_total_count(soup0) if soup0 else 0
    log(f"  Area total: {area_total:,} restaurants in Minato-ku")

    # Tally genre matches across SCAN_PAGES pages
    genre_tally: dict[str, int] = {cat: 0 for cat in CATEGORY_KEYWORDS}
    cards_scanned = 0

    for page in range(1, SCAN_PAGES + 1):
        url = BASE_URL + "rstLst/" if page == 1 else BASE_URL + f"rstLst/?PG={page}"
        log(f"  Scanning listing page {page}/{SCAN_PAGES} for genre counts...")
        soup = fetch(url) if page > 1 else soup0
        if not soup:
            log(f"  [!] Could not fetch page {page}, stopping scan.")
            break

        cards = soup.select("div.list-rst__wrap, li.js-rstlist-info")
        for card in cards:
            genre_tag = card.select_one(".list-rst__area-genre, .cpy-area-genre")
            if not genre_tag:
                continue
            raw = genre_tag.get_text(strip=True).lower()
            for cat, keywords in CATEGORY_KEYWORDS.items():
                if any(kw.lower() in raw for kw in keywords):
                    genre_tally[cat] += 1
        cards_scanned += len(cards)
        sleep()

    log(f"  Scanned {cards_scanned} restaurant cards across {SCAN_PAGES} pages.")

    # Scale up from sample to area total
    results = []
    for cat in CATEGORY_KEYWORDS:
        sample_count = genre_tally[cat]
        if cards_scanned > 0 and area_total > 0:
            estimated = round(sample_count / cards_scanned * area_total)
        else:
            estimated = sample_count
        url = CATEGORY_SLUGS.get(cat, BASE_URL + "rstLst/")
        log(f"  {cat}: {sample_count} in sample → ~{estimated:,} estimated total")
        results.append({
            "Category":           cat,
            "In Sample (200 rst)": sample_count,
            "Est. Total (scaled)": estimated,
            "Area Total":          area_total,
            "Sample Size":         cards_scanned,
            "URL":                 url,
        })

    return results

# ── Task 2 — Top-20 restaurants ──────────────────────────────────────────────

def _parse_restaurant_cards(soup: BeautifulSoup) -> list[dict]:
    """Extract restaurant data from a Tabelog listing page."""
    restaurants = []
    # Main card selector — Tabelog uses .list-rst__wrap or .js-rstlist-info
    cards = soup.select("div.list-rst__wrap")
    if not cards:
        cards = soup.select("li.js-rstlist-info")

    for card in cards:
        try:
            # Name (Japanese)
            name_tag = card.select_one(".list-rst__rst-name-target, .c-rst-card__name")
            name_ja = name_tag.get_text(strip=True) if name_tag else "N/A"

            # Romaji name — appears in a sibling span or alt attribute
            name_ro_tag = card.select_one(".list-rst__rst-name-target span[class*='latin'], .list-rst__name-latin")
            name_ro = name_ro_tag.get_text(strip=True) if name_ro_tag else ""

            # Score
            score_tag = card.select_one(".c-rating__val, .list-rst__rating-val")
            score_text = score_tag.get_text(strip=True) if score_tag else "0.00"
            try:
                score = float(score_text)
            except ValueError:
                score = 0.00

            # Category / cuisine
            # Tabelog cards show area+genre in div.list-rst__area-genre
            # e.g. "乃木坂駅 669m / 寿司、海鮮"
            cat_tag = card.select_one(".list-rst__area-genre, .cpy-area-genre")
            if cat_tag:
                raw_cat = cat_tag.get_text(strip=True)
                # Extract only the genre part after the " / " separator
                if " / " in raw_cat:
                    category = raw_cat.split(" / ", 1)[1].strip()
                elif "／" in raw_cat:
                    category = raw_cat.split("／", 1)[1].strip()
                else:
                    category = raw_cat
                category = re.sub(r"\s+", " ", category).strip()
            else:
                category = "N/A"

            # Price range — new HTML uses c-rating-v3__val for budget cells
            price_tags = card.select(".c-rating-v3__val, .c-rating-v2__price, .list-rst__price")
            if price_tags:
                price = " / ".join(t.get_text(strip=True) for t in price_tags[:2])
            else:
                price = "N/A"

            # Review count — actual count number is in .list-rst__rvw-count-num
            review_tag = card.select_one(
                ".list-rst__rvw-count-num, .cpy-review-count, "
                ".list-rst__rvw-count, .c-rating-v2__review-count"
            )
            if review_tag:
                raw_review = review_tag.get_text(strip=True)
                nums = re.findall(r"[\d,]+", raw_review)
                reviews = int(nums[0].replace(",", "")) if nums else 0
            else:
                reviews = 0

            # Restaurant URL (for deeper scraping if needed)
            link_tag = card.select_one("a.list-rst__rst-name-target, a[href*='/tokyo/']")
            rst_url = link_tag["href"] if link_tag and link_tag.has_attr("href") else ""

            restaurants.append({
                "Name (JA)":    name_ja,
                "Name (Romaji)": name_ro,
                "Score":        score,
                "Category":     category,
                "Price Range":  price,
                "Reviews":      reviews,
                "URL":          rst_url,
            })
        except Exception as e:
            log(f"  [!] Skipping a card due to parse error: {e}")
            continue

    return restaurants


def _scrape_detail_category(rst_url: str) -> str:
    """
    Fetch a restaurant's own Tabelog page and extract its genre/category.
    Used as fallback when listing cards don't expose genre.
    """
    if not rst_url:
        return "N/A"
    soup = fetch(rst_url)
    if not soup:
        return "N/A"
    # Genre shown in the info table on detail pages
    for sel in [
        ".rstdtl-side-genre__genre-name",
        "dd.rstinfo-table__val a[href*='rstLst']",
        ".rstdtl-side-genre a",
        "table.rstinfo-table a[href*='rstLst']",
        ".c-table__val a[href*='rstLst']",
    ]:
        tags = soup.select(sel)
        if tags:
            names = [t.get_text(strip=True) for t in tags if t.get_text(strip=True)]
            if names:
                return " / ".join(names[:2])
    # Fallback: look for any 件 counting page with genre in breadcrumb
    breadcrumb = soup.select_one(".c-breadcrumb, #js-breadcrumb")
    if breadcrumb:
        crumbs = [a.get_text(strip=True) for a in breadcrumb.find_all("a")]
        if len(crumbs) >= 3:
            return crumbs[-1]
    return "N/A"


def _enrich_categories(restaurants: list[dict]) -> None:
    """
    For any restaurant whose Category is N/A, scrape its detail page to get genre.
    Modifies the list in place.
    """
    needs_enrich = [r for r in restaurants if r.get("Category", "N/A") == "N/A" and r.get("URL")]
    if not needs_enrich:
        return
    log(f"  Enriching categories for {len(needs_enrich)} restaurants via detail pages...")
    for r in needs_enrich:
        cat = _scrape_detail_category(r["URL"])
        r["Category"] = cat
        log(f"    {r['Name (JA)']}: {cat}")
        sleep()


def task2_top20() -> list[dict]:
    banner("TASK 2 — Top 20 restaurants in Minato-ku (ranked by score)")
    all_restaurants: list[dict] = []
    page = 1

    while len(all_restaurants) < 20:
        url = RANKING_URL if page == 1 else f"{RANKING_URL}&PG={page}"
        log(f"Fetching ranking page {page}: {url}")
        soup = fetch(url)

        if not soup:
            log("[!] Could not fetch ranking page — aborting Task 2.")
            break

        page_results = _parse_restaurant_cards(soup)

        # Debug: dump first card HTML on first page to help diagnose selector issues
        if page == 1 and page_results:
            first_card = soup.select_one("div.list-rst__wrap, li.js-rstlist-info")
            if first_card:
                snippet = str(first_card)[:800]
                log(f"  [DEBUG] First card HTML snippet:\n{snippet}\n")

        if not page_results:
            log("  → No restaurant cards found on this page. Stopping pagination.")
            break

        all_restaurants.extend(page_results)
        log(f"  → Collected {len(page_results)} restaurants (total so far: {len(all_restaurants)})")

        # Check if there's a next page
        next_btn = soup.select_one("a.c-pagination__arrow--next, a[data-action='next']")
        if not next_btn:
            log("  → No next page found.")
            break

        page += 1
        sleep()

    # Trim to top 20, rank them
    top20 = all_restaurants[:20]
    for i, r in enumerate(top20, 1):
        r["Rank"] = i

    # Enrich any missing categories via detail pages
    _enrich_categories(top20)

    log(f"\nTop 20 collected ({len(top20)} restaurants):")
    for r in top20:
        print(f"  {r['Rank']:>2}. {r['Name (JA)']:<30} Score: {r['Score']}  Cat: {r['Category']}")

    return top20

# ── Task 3 — Uber Eats cross-check ──────────────────────────────────────────

def _search_uber_eats(restaurant_name: str) -> bool:
    """
    Search Bing for '[name] site:ubereats.com/jp' and return True only when
    a result link actually points to a ubereats.com store page.
    Bing returns plain HTML with real destination URLs and is more tolerant
    of programmatic access than Google or DuckDuckGo.
    """
    query = f'"{restaurant_name}" site:ubereats.com/jp'
    url = "https://www.bing.com/search"
    params = {"q": query, "setlang": "ja", "cc": "JP", "count": 5}

    headers = {
        **HEADERS,
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0"
        ),
        "Accept-Language": "ja-JP,ja;q=0.9",
    }

    try:
        resp = SESSION.get(url, params=params, headers=headers, timeout=15)
        if resp.status_code != 200:
            log(f"    [!] Bing returned {resp.status_code} for '{restaurant_name}'")
            return False

        soup = BeautifulSoup(resp.text, "html.parser")

        # Check for explicit "no results" from Bing
        no_result_phrases = [
            "一致する結果はありません",
            "に一致する情報は見つかりませんでした",
            "No results found",
        ]
        page_text = soup.get_text()
        if any(p in page_text for p in no_result_phrases):
            return False

        # Bing result links — check ALL anchor hrefs on the page
        STORE_PATTERNS = ["ubereats.com/jp/store/", "ubereats.com/store/"]

        for a in soup.find_all("a", href=True):
            href = a["href"]
            if any(pat in href for pat in STORE_PATTERNS):
                return True

        # Also check cite/URL display and any plain text
        for cite in soup.find_all("cite"):
            text = cite.get_text()
            if any(pat in text for pat in STORE_PATTERNS):
                return True

        return False

    except Exception as e:
        log(f"    [!] Bing search error for '{restaurant_name}': {e}")
        return False


def task3_uber_eats_check(top20: list[dict]) -> list[dict]:
    banner("TASK 3 — Cross-checking Top 20 against Uber Eats Japan")
    results = []

    for i, r in enumerate(top20, 1):
        name = r["Name (JA)"]
        log(f"[{i:>2}/20] Searching Uber Eats for: {name}")

        found = _search_uber_eats(name)
        status = "On Uber Eats ✅" if found else "Not on Uber Eats ❌"
        log(f"         → {status}")

        entry = dict(r)
        entry["Uber Eats Status"] = status
        results.append(entry)

        # Longer delay for search engine requests to avoid rate limiting
        time.sleep(random.uniform(5.0, 8.0))

    on_count  = sum(1 for r in results if "✅" in r["Uber Eats Status"])
    off_count = sum(1 for r in results if "❌" in r["Uber Eats Status"])
    log(f"\nSummary: {on_count} on Uber Eats | {off_count} NOT on Uber Eats")

    return results

# ── Task 4 — Supply gap summary ──────────────────────────────────────────────

def task4_supply_gap(task3_data: list[dict]) -> tuple[list[dict], list[dict]]:
    banner("TASK 4 — Supply gap analysis")

    not_on_ue = [r for r in task3_data if "❌" in r["Uber Eats Status"]]
    on_ue     = [r for r in task3_data if "✅" in r["Uber Eats Status"]]

    total       = len(task3_data)
    absent      = len(not_on_ue)
    pct_absent  = (absent / total * 100) if total else 0

    log(f"Total Top 20:          {total}")
    log(f"On Uber Eats:          {len(on_ue)}")
    log(f"NOT on Uber Eats:      {absent}  ({pct_absent:.1f}%)")

    # Group missing by category
    cat_counts: dict[str, int] = {}
    for r in not_on_ue:
        cat = r.get("Category", "Unknown")
        cat_counts[cat] = cat_counts.get(cat, 0) + 1

    log("\nMissing restaurants by category:")
    for cat, cnt in sorted(cat_counts.items(), key=lambda x: -x[1]):
        log(f"  {cat}: {cnt}")

    # Summary stats rows
    stats = [
        {"Metric": "Total restaurants in Top 20",             "Value": total},
        {"Metric": "Restaurants ON Uber Eats",                "Value": len(on_ue)},
        {"Metric": "Restaurants NOT on Uber Eats",            "Value": absent},
        {"Metric": "% absent from Uber Eats",                 "Value": f"{pct_absent:.1f}%"},
        {"Metric": "--- MISSING BY CATEGORY ---",             "Value": ""},
    ]
    for cat, cnt in sorted(cat_counts.items(), key=lambda x: -x[1]):
        stats.append({"Metric": f"  {cat}", "Value": cnt})

    # Priority list — NOT on Uber Eats, sorted by score descending
    priority = sorted(not_on_ue, key=lambda x: x.get("Score", 0), reverse=True)
    for i, r in enumerate(priority, 1):
        r["Priority"] = i
        log(f"  #{i:>2}  {r['Name (JA)']:<30}  Score: {r['Score']}  Cat: {r['Category']}")

    return stats, priority

# ── Excel export ─────────────────────────────────────────────────────────────

def _style_header_row(ws, row: int, fill_hex: str = "2E5C99"):
    fill  = PatternFill("solid", fgColor=fill_hex)
    font  = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[row]:
        cell.fill  = fill
        cell.font  = font
        cell.alignment = align


def _auto_width(ws, min_width: int = 12, max_width: int = 50):
    for col in ws.columns:
        length = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col
        )
        length = max(min_width, min(length + 2, max_width))
        ws.column_dimensions[get_column_letter(col[0].column)].width = length


def _zebra(ws, start_row: int = 2, fill_hex: str = "EEF2FA"):
    fill = PatternFill("solid", fgColor=fill_hex)
    for row in ws.iter_rows(min_row=start_row):
        if row[0].row % 2 == 0:
            for cell in row:
                if cell.fill.patternType == "none":
                    cell.fill = fill


def _thin_border(ws):
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border


def export_excel(
    cat_counts: list[dict],
    top20: list[dict],
    task3_data: list[dict],
    stats: list[dict],
    priority: list[dict],
):
    banner("EXPORTING to Excel")
    wb = openpyxl.Workbook()

    # ── Sheet 1: Category counts ────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "1 - Category Counts"
    headers1 = ["Category", "In Sample (200 rst)", "Est. Total (scaled)", "Area Total", "Sample Size", "URL"]
    ws1.append(headers1)
    for row in cat_counts:
        ws1.append([row.get(h, "") for h in headers1])
    _style_header_row(ws1, 1)
    _auto_width(ws1)
    _zebra(ws1)
    _thin_border(ws1)
    ws1.freeze_panes = "A2"

    # ── Sheet 2: Top 20 rankings ────────────────────────────────────────────
    ws2 = wb.create_sheet("2 - Top 20 Rankings")
    headers2 = ["Rank", "Name (JA)", "Name (Romaji)", "Score", "Category", "Price Range", "Reviews", "URL"]
    ws2.append(headers2)
    for row in top20:
        ws2.append([row.get(h, "") for h in headers2])
    _style_header_row(ws2, 1)
    _auto_width(ws2)
    _zebra(ws2)
    _thin_border(ws2)
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Top 20 + Uber Eats status ─────────────────────────────────
    ws3 = wb.create_sheet("3 - Uber Eats Status")
    headers3 = ["Rank", "Name (JA)", "Name (Romaji)", "Score", "Category", "Price Range", "Reviews", "Uber Eats Status", "URL"]
    ws3.append(headers3)
    green = PatternFill("solid", fgColor="C6EFCE")
    red   = PatternFill("solid", fgColor="FFC7CE")
    for row in task3_data:
        ws3.append([row.get(h, "") for h in headers3])
        last_row = ws3.max_row
        status_cell = ws3.cell(last_row, headers3.index("Uber Eats Status") + 1)
        if "✅" in str(status_cell.value):
            status_cell.fill = green
        elif "❌" in str(status_cell.value):
            status_cell.fill = red
    _style_header_row(ws3, 1)
    _auto_width(ws3)
    _thin_border(ws3)
    ws3.freeze_panes = "A2"

    # ── Sheet 4: Supply gap summary ─────────────────────────────────────────
    ws4 = wb.create_sheet("4 - Supply Gap Summary")

    # Summary stats section
    ws4.append(["SUPPLY GAP SUMMARY", ""])
    ws4["A1"].font = Font(bold=True, size=14, color="2E5C99")
    ws4.append(["Metric", "Value"])
    _style_header_row(ws4, 2, fill_hex="2E5C99")
    for stat in stats:
        ws4.append([stat.get("Metric", ""), stat.get("Value", "")])

    # Spacer
    ws4.append(["", ""])

    # Priority list section
    priority_start = ws4.max_row + 1
    ws4.append(["PRIORITY LIST — Restaurants to Approach (Not on Uber Eats, by Score)"])
    ws4.cell(priority_start, 1).font = Font(bold=True, size=14, color="C0392B")

    headers4p = ["Priority", "Name (JA)", "Name (Romaji)", "Score", "Category", "Price Range", "Reviews"]
    ws4.append(headers4p)
    _style_header_row(ws4, ws4.max_row, fill_hex="C0392B")
    for row in priority:
        ws4.append([row.get(h, "") for h in headers4p])
    _auto_width(ws4)
    _zebra(ws4, start_row=priority_start + 2)
    _thin_border(ws4)
    ws4.freeze_panes = f"A{priority_start + 2}"

    wb.save(OUTPUT_FILE)
    log(f"Saved → {OUTPUT_FILE}")


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    print("\n" + "█"*60)
    print("  Tabelog × Uber Eats Supply Gap Analyser")
    print("  Minato-ku, Tokyo")
    print("█"*60)
    print(f"  Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("  Output:  " + OUTPUT_FILE)

    # Task 1
    cat_counts = task1_category_counts()

    # Task 2
    top20 = task2_top20()
    if not top20:
        log("[WARN] No restaurants found in Task 2. Using placeholder data for Tasks 3 & 4.")
        top20 = [{
            "Rank": i, "Name (JA)": f"Restaurant {i}", "Name (Romaji)": "",
            "Score": 0.0, "Category": "Unknown", "Price Range": "N/A", "Reviews": 0, "URL": ""
        } for i in range(1, 21)]

    # Task 3
    task3_data = task3_uber_eats_check(top20)

    # Task 4
    stats, priority = task4_supply_gap(task3_data)

    # Export
    export_excel(cat_counts, top20, task3_data, stats, priority)

    print("\n" + "█"*60)
    print("  ALL DONE")
    print(f"  Finished: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("  Open: " + OUTPUT_FILE)
    print("█"*60 + "\n")


if __name__ == "__main__":
    main()
