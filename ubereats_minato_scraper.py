#!/usr/bin/env python3
"""
Uber Eats Japan — 港区 (Minato-ku) レストラン カテゴリ別調査
Playwright でカテゴリ検索ページを巡回し、
各カテゴリのレストラン数と店名一覧を取得する。

Output: ubereats_minato_categories.xlsx
"""

from __future__ import annotations
import json
import base64
import urllib.parse
import time
import sys
import re
from collections import defaultdict
from datetime import datetime

MISSING = []
try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
except ImportError:
    MISSING.append("playwright")
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    MISSING.append("openpyxl")

if MISSING:
    print(f"[ERROR] Missing: {', '.join(MISSING)} — pip install {' '.join(MISSING)}")
    sys.exit(1)

# ── config ───────────────────────────────────────────────────────────────────
MINATO_LAT  = 35.6580
MINATO_LNG  = 139.7454
MINATO_ADDR = "東京都港区"

_loc = json.dumps({"address": MINATO_ADDR, "latitude": MINATO_LAT, "longitude": MINATO_LNG})
PL_PARAM = urllib.parse.quote(base64.b64encode(_loc.encode()).decode())

OUTPUT_FILE = "ubereats_minato_categories.xlsx"

# カテゴリ一覧 — Uber Eats Japan の検索キーワード
CATEGORIES = [
    ("寿司 (Sushi)",               "寿司"),
    ("ラーメン (Ramen)",            "ラーメン"),
    ("ピザ (Pizza)",               "ピザ"),
    ("イタリアン (Italian)",        "イタリアン"),
    ("焼肉 (Yakiniku)",            "焼肉"),
    ("天ぷら (Tempura)",           "天ぷら"),
    ("フレンチ (French)",           "フレンチ"),
    ("中華料理 (Chinese)",          "中華料理"),
    ("バーガー (Burger)",           "ハンバーガー"),
    ("カレー (Curry)",             "カレー"),
    ("和食 (Japanese)",            "和食"),
    ("丼 (Rice Bowl)",             "丼"),
    ("焼き鳥 (Yakitori)",          "焼き鳥"),
    ("ステーキ (Steak)",           "ステーキ"),
    ("タイ料理 (Thai)",            "タイ料理"),
    ("インド料理 (Indian)",         "インド料理"),
    ("韓国料理 (Korean)",           "韓国料理"),
    ("サンドイッチ (Sandwich)",     "サンドイッチ"),
    ("スイーツ・カフェ (Sweets)",    "スイーツ"),
    ("そば・うどん (Noodles)",      "そば"),
    ("から揚げ (Karaage)",         "から揚げ"),
    ("居酒屋 (Izakaya)",           "居酒屋"),
    ("ベトナム料理 (Vietnamese)",    "ベトナム料理"),
    ("メキシカン (Mexican)",        "メキシカン"),
    ("ヘルシー (Healthy)",          "ヘルシー料理"),
]


def log(msg: str):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")


def build_search_url(query: str) -> str:
    q = urllib.parse.quote(query)
    return f"https://www.ubereats.com/jp/search?q={q}&pl={PL_PARAM}&diningMode=DELIVERY"


def scroll_and_collect(page, max_scrolls: int = 30) -> list[str]:
    """
    Scroll through a search result page and collect all restaurant names.
    Returns deduplicated list of names.
    """
    seen: set[str] = set()
    names: list[str] = []
    stale = 0
    prev_count = 0

    for _ in range(max_scrolls):
        # Collect current DOM
        cards = page.evaluate("""
        () => {
            const out = [];
            for (const a of document.querySelectorAll('[data-testid="store-card"]')) {
                const h = a.querySelector('h3, h2');
                if (h) out.push(h.innerText.trim());
            }
            return out;
        }
        """)
        for name in cards:
            if name and name not in seen:
                seen.add(name)
                names.append(name)

        # Scroll
        page.evaluate("window.scrollBy(0, 1000)")
        time.sleep(1.5)

        if len(names) == prev_count:
            stale += 1
            if stale >= 4:
                break
        else:
            stale = 0
        prev_count = len(names)

    return names


def scrape_category(page, label: str, query: str) -> list[str]:
    """Navigate to category search and return list of restaurant names."""
    url = build_search_url(query)
    log(f"  [{label}] → {url[:90]}")
    try:
        page.goto(url, timeout=25000, wait_until="domcontentloaded")
    except PWTimeout:
        log(f"    [!] Timeout navigating — continuing")

    # Wait for first result
    try:
        page.wait_for_selector('[data-testid="store-card"]', timeout=12000)
    except PWTimeout:
        log(f"    [!] No store cards appeared for '{query}'")
        return []

    time.sleep(2)
    names = scroll_and_collect(page, max_scrolls=40)
    log(f"    → {len(names)} restaurants found")
    return names


def run_all() -> tuple[dict[str, list[str]], list[str]]:
    """
    Run through all categories. Also collect the full feed for total count.
    Returns (category_map, all_feed_names).
    """
    category_results: dict[str, list[str]] = {}

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        context = browser.new_context(
            locale="ja-JP",
            timezone_id="Asia/Tokyo",
            viewport={"width": 1280, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
        )
        page = context.new_page()

        # ── Step 1: get overall feed count ───────────────────────────────────
        feed_url = f"https://www.ubereats.com/jp/feed?diningMode=DELIVERY&pl={PL_PARAM}"
        log(f"Getting total feed (all Minato-ku restaurants)...")
        try:
            page.goto(feed_url, timeout=25000, wait_until="domcontentloaded")
            page.wait_for_selector('[data-testid="store-card"]', timeout=15000)
        except PWTimeout:
            log("  [!] Feed timeout")

        time.sleep(3)
        all_feed_names = scroll_and_collect(page, max_scrolls=60)
        log(f"  → Feed total: {len(all_feed_names)} restaurants")

        # ── Step 2: per-category search ───────────────────────────────────────
        log(f"\nSearching {len(CATEGORIES)} categories...")
        for label, query in CATEGORIES:
            names = scrape_category(page, label, query)
            category_results[label] = names
            time.sleep(2)

        browser.close()

    return category_results, all_feed_names


# ── Excel export ─────────────────────────────────────────────────────────────

def export_excel(
    category_results: dict[str, list[str]],
    all_feed_names: list[str],
):
    wb = openpyxl.Workbook()
    BLUE = "1A56DB"
    HDR_FILL = PatternFill("solid", fgColor=BLUE)
    HDR_FONT = Font(bold=True, color="FFFFFF")
    HDR_ALIGN = Alignment(horizontal="center")
    ZEBRA = PatternFill("solid", fgColor="EEF4FF")

    def style_header(ws, row_num):
        for cell in ws[row_num]:
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = HDR_ALIGN

    # ── Sheet 1: Category summary ────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "カテゴリ別件数"

    ws1["A1"] = "Uber Eats 港区 — カテゴリ別レストラン数"
    ws1["A1"].font = Font(bold=True, size=14, color=BLUE)
    ws1["A2"] = f"取得日時: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws1["A3"] = f"フィードの総レストラン数: {len(all_feed_names)} 件"
    ws1["A4"] = "※ カテゴリ検索はUber Eats Japanの検索機能を利用（地域:港区）"
    ws1.append([])
    ws1.append(["カテゴリ", "件数", "Tabelog 港区 推定総数 (参考)", "Uber Eats カバー率 (参考)"])
    style_header(ws1, 6)

    TABELOG_ESTIMATES = {
        "寿司 (Sushi)":         326,
        "ラーメン (Ramen)":      21,
        "ピザ (Pizza)":          42,
        "イタリアン (Italian)":  262,
        "焼肉 (Yakiniku)":      231,
        "天ぷら (Tempura)":      21,
        "フレンチ (French)":    105,
    }

    sorted_cats = sorted(category_results.items(), key=lambda x: -len(x[1]))
    for i, (label, names) in enumerate(sorted_cats, 7):
        cnt = len(names)
        tabelog_est = TABELOG_ESTIMATES.get(label, "—")
        coverage = f"{cnt/tabelog_est*100:.0f}%" if isinstance(tabelog_est, int) and tabelog_est > 0 else "—"
        ws1.append([label, cnt, tabelog_est, coverage])
        if i % 2 == 0:
            for cell in ws1[i]:
                cell.fill = ZEBRA

    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 8
    ws1.column_dimensions["C"].width = 28
    ws1.column_dimensions["D"].width = 22
    ws1.freeze_panes = "A7"

    # ── Sheet 2: Full feed list ──────────────────────────────────────────────
    ws2 = wb.create_sheet("全レストランリスト")
    ws2.append(["#", "レストラン名 (フィード全体)"])
    style_header(ws2, 1)
    for i, name in enumerate(all_feed_names, 1):
        ws2.append([i, name])
        if i % 2 == 0:
            ws2.cell(i + 1, 1).fill = ZEBRA
            ws2.cell(i + 1, 2).fill = ZEBRA
    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 50
    ws2.freeze_panes = "A2"

    # ── Per-category sheets ──────────────────────────────────────────────────
    for label, names in sorted_cats:
        safe_title = re.sub(r'[\\/*?:\[\]]', '', label)[:28]
        ws = wb.create_sheet(safe_title)
        ws.append(["#", f"{label} — {len(names)} 件"])
        style_header(ws, 1)
        for i, name in enumerate(names, 1):
            ws.append([i, name])
            if i % 2 == 0:
                ws.cell(i + 1, 1).fill = ZEBRA
                ws.cell(i + 1, 2).fill = ZEBRA
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 50
        ws.freeze_panes = "A2"

    wb.save(OUTPUT_FILE)
    log(f"Saved → {OUTPUT_FILE}")


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    print("\n" + "█" * 60)
    print("  Uber Eats Japan — 港区 カテゴリ別レストラン調査")
    print("█" * 60)
    print(f"  開始: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

    category_results, all_feed_names = run_all()

    print(f"\n{'='*60}")
    print("  結果サマリー")
    print(f"{'='*60}")
    print(f"  フィード総数: {len(all_feed_names)} 件\n")

    sorted_cats = sorted(category_results.items(), key=lambda x: -len(x[1]))
    for label, names in sorted_cats:
        bar = "█" * min(len(names) // 3, 40)
        print(f"  {label:<35} {len(names):>4} 件  {bar}")

    export_excel(category_results, all_feed_names)

    print(f"\n  完了: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  出力: {OUTPUT_FILE}\n")


if __name__ == "__main__":
    main()
